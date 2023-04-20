from openpyxl import load_workbook

# Excel dosyasını açar
wb = load_workbook('ENGR 102 Student List.xlsx')

# Aktif sayfayı seçer
ws = wb.active

# ----------------- Bölüm Seçimi -----------------
# combo_box_items listesini tanımlar. Yani bolumlerin listesi
combo_box_items = []

# Ilk satırı haric D sütunundaki tüm hücreleri ekler
for cell in ws['D'][1:]:
    combo_box_items.append(cell.value)

# combo_box_items listesindeki tekrar eden elemanları siler
combo_box_items = list(set(combo_box_items))

# combo_box_items listesindeki elemanları alfabetik sıraya koyar
combo_box_items.sort()

# combo_box_items listesindeki elemanları ekrana yazdırır
for items in combo_box_items:
    print(items)

user_choice_section = int(input("Bolum Seciniz:"))
user_choice_section = combo_box_items[user_choice_section - 1]
print(user_choice_section)


# min row 2. satirdan baslar cunku basliklar olmayacak, min col 1. sutundan baslayip, max col 4. sutuna kadar döngüye alir

# ----------------- Öğrenci Seçimi Function -----------------


def section_students(section: str):
    # min row 2. satirdan baslar cunku basliklar olmayacak, min col 1. sutundan baslayip, max col 4. sutuna kadar döngüye alir
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=4):
        cell = row[3]  # 4. sutun
        if cell.value == section:  # 4. sutunun degeri bolumun degerine esitse
            # 2. sutunun degerini bosluklara gore ayirir
            name_surname = row[1].value.split()
            # 2. sutunun ilk elemani soyadi, ikinci elemani adi, 1. sutunun degeri numarasi
            print(name_surname[1], name_surname[0], row[0].value)


section_students(section=user_choice_section)


