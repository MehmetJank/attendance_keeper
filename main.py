# Description: Attendance Keeper App for school project
# Author: Mehmet Can Kaya

from tkinter import (
    Tk,
    ttk,
    Frame,
    Entry,
    Button,
    END,
    RIGHT,
    Label,
    Listbox,
    StringVar,
    filedialog,
    messagebox,
)
from openpyxl import load_workbook
from xlwt import Workbook


class ImportExportExcel:
    def __init__(self):
        self.attented_students = []

    def open_file_dialog(self):
        """Open file dialog and get file path"""
        self.file_path = filedialog.askopenfilename(
            title="Import Excel File",
            filetypes=(("Excel files", "*.xlsx*"), ("all files", "*.*")),
        )
        self.wb = load_workbook(self.file_path)
        self.ws = self.wb.active

    def get_combo_box_items(self):
        """Get combo box items from excel file"""
        self.combo_box_items = []
        if hasattr(self, "ws"):
            self.combo_box_items = list(set(cell.value for cell in self.ws["D"][1:]))
            self.combo_box_items.sort()
            return self.combo_box_items

    def get_section(self, index: int):
        """Get section from excel file"""
        combo_box_items = self.get_combo_box_items()
        if combo_box_items is None or index >= len(combo_box_items):
            return []
        section = combo_box_items[index]
        return (
            f"{row[1].value.rsplit(' ', 1)[-1]}, {row[1].value.rsplit(' ', 1)[0]}, {row[0].value}"
            if " " in row[1].value
            else f"{row[1].value}, {row[0].value}"
            for row in self.ws.iter_rows(min_row=2, min_col=1, max_col=4)
            if row[3].value == section
        )

    def get_attended_students(self, item: str = None, process: str = None):
        """Add, remove or clear attended students"""
        if process == "append":
            self.attented_students.append(item)
        elif process == "remove":
            self.attented_students.remove(item)
        elif process == "clear":
            self.attented_students.clear()

    def export_type(self, file_type: str, file_name: str):
        if file_type == "txt":
            with open(file_name, "w", encoding="utf-8") as f:
                for student in self.attented_students:
                    student_info = student.split(", ")
                    student_id = student_info[-1]
                    name = ", ".join(student_info[:-1])
                    department = self.ws[f"C{student_id}"].value
                    f.write(f"{name}, {department}\n")
                messagebox.showinfo("Success", "Exported to txt file successfully!")
        elif file_type == "xls":
            workbook = Workbook()
            worksheet = workbook.add_sheet("Attendance")
            row_num = 0
            # Create ID ,Name and Department columns
            columns = ["ID", "Name", "Department"]
            for col_num, column_title in enumerate(columns):
                worksheet.write(row_num, col_num, column_title)
            row_num += 1
            for student in self.attented_students:
                student_info = student.split(", ")
                student_id = student_info[-1]
                name = ", ".join(student_info[:-1])
                department = self.ws[f"C{student_id}"].value
                # ID, Name ve Department sütunlarına verileri yaz
                worksheet.write(row_num, 0, student_id)
                worksheet.write(row_num, 1, name)
                worksheet.write(row_num, 2, department)
                row_num += 1
            workbook.save(file_name)
            messagebox.showinfo("Success", "Exported to xls file successfully!")
        elif file_type == "csv":
            messagebox.showerror("Error", "CSV File type is not supported!")
            raise BaseException("CSV File type is not supported!")


class AttendanceKeeperApp(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent)
        self.parent = parent
        self.import_export_excel = ImportExportExcel()
        self.selected_section = StringVar()
        self.initUI()

    # Custom Button
    def custom_button(
        self,
        text: str,
        row: int,
        column: int,
        sticky: str,
        font: str,
        command: object,
        width: int = 0,
        height: int = 0,
        columnspan: int = 1,
        rowspan: int = 1,
    ):
        button = Button(
            text=text, width=width, height=height, command=command, font=font
        )
        button.grid(
            row=row,
            column=column,
            columnspan=columnspan,
            rowspan=rowspan,
            padx=3,
            pady=3,
            sticky=sticky,
        )

    # Custom Text Label
    def custom_text_label(
        self,
        text: str,
        row: int,
        column: int,
        sticky: str,
        font: str,
        width: int = 0,
        height: int = 0,
        columnspan: int = 1,
        rowspan: int = 1,
    ):
        label = Label(
            text=text,
            width=width,
            height=height,
            font=font,
        )
        label.grid(
            row=row,
            column=column,
            columnspan=columnspan,
            rowspan=rowspan,
            padx=3,
            pady=3,
            sticky=sticky,
        )

    # Custom Listbox
    def custom_listbox(
        self,
        row: int,
        column: int,
        list_values: list = [],
        listbox: Listbox = None,
        bind: object = None,
    ):
        if listbox is None:
            listbox = Listbox(width=40, height=5, selectmode="multiple")
            listbox.grid(
                row=row,
                column=column,
                columnspan=2,
                rowspan=3,
                padx=3,
                pady=3,
                sticky="nswe",
            )
        else:
            listbox.delete(0, END)

        for item in list_values:
            listbox.insert(END, item)

        listbox.bind("<<ListboxSelect>>", lambda event: bind)

        return listbox

    # Custom Combo Box
    def custom_combo_box(
        self,
        row: int,
        column: int,
        width: int,
        height: int,
        textvariable: str = None,
    ):
        combo = ttk.Combobox(width=width, height=height, textvariable=textvariable)
        combo.grid(
            row=row,
            column=column,
            columnspan=1,
            rowspan=1,
            padx=3,
            pady=3,
            sticky="nswe",
        )

        return combo

    def update_listbox(self, index: int = 0, side: int = 0):
        """Updates the listbox with the selected combo box item."""
        if side == 0:  # side 0 is left listbox
            self.list_values_left = []
            self.list_values_left = self.import_export_excel.get_section(index)
            self.list_left = self.custom_listbox(
                row=3, column=0, list_values=self.list_values_left
            )
            if hasattr(self, "list_right"):
                self.list_right.delete(0, END)
                self.import_export_excel.get_attended_students(process="clear")
        else:  # side 1 is right listbox
            self.list_values_right = []
            self.list_right = self.custom_listbox(
                row=3, column=3, list_values=self.list_values_right
            )

    def move_items_between_listboxes(self, source_listbox, target_listbox):
        """Moves the selected items between listboxes"""
        selected_items = source_listbox.curselection()
        if selected_items == ():
            messagebox.showerror("Error", "No item selected!")
            return
        selected_items = sorted(selected_items, reverse=True)
        for item_in_list in selected_items:
            item = source_listbox.get(item_in_list)
            if target_listbox == self.list_right:
                self.import_export_excel.get_attended_students(
                    item=item, process="append"
                )
            else:
                self.import_export_excel.get_attended_students(
                    item=item, process="remove"
                )
                source_listbox.delete(item_in_list)

            target_listbox.insert(END, item)

    def update_combo_box(self, index: int = 0):
        """Updates the combo box with the selected file."""
        self.section_combo_box_items = self.import_export_excel.get_combo_box_items()
        # if combo box empty then add a default value
        if self.section_combo_box_items == None:
            self.section_combo_box_items = ["Import a file first"]
        self.section_combo_box = self.custom_combo_box(
            row=3,
            column=2,
            width=15,
            height=5,
            textvariable=self.selected_section,
        )
        self.section_combo_box["values"] = self.section_combo_box_items
        self.section_combo_box.current(0)
        self.section_combo_box.bind(
            "<<ComboboxSelected>>",
            lambda event: self.update_listbox(
                index=self.section_combo_box.current(), side=0
            ),
        )
        self.update_listbox(index=self.section_combo_box.current(), side=0)

    def import_button(self):
        """This function open file explorer and get the excel file and call import_file function"""
        self.import_export_excel.open_file_dialog(),
        self.update_combo_box()
        self.combo_box_items = self.import_export_excel.get_combo_box_items()

    def export_button(self):
        """This fuction get export type and file name from user and export the excel file and call export_type function"""
        week_value = self.week_input.get()
        section_value = self.selected_section.get()
        file_type = self.file_type_combo_box.get()
        if week_value == "" or section_value == "":
            messagebox.showerror("Error", "Week and Section must be selected!")
            return
        file_name = section_value + " Week " + week_value + "." + file_type
        self.import_export_excel.export_type(file_type, file_name)

    def initUI(self):
        # Boyutlandırma
        for i in range(5):
            self.master.columnconfigure(i, weight=1)
        for i in range(7):
            self.master.rowconfigure(i, weight=1)

        # Window title
        self.parent.title("Attendance Keeper")

        # Row 0
        # Application title
        self.custom_text_label(
            text="AttandanceKeeper v1.0",
            row=0,
            column=0,
            font="Arial 24 bold",
            sticky="nswe",
            columnspan=5,
        )

        # Row 1
        # Select Student list Excel File text label
        self.custom_text_label(
            text="Select Student list Excel File:",
            row=1,
            column=0,
            font="Arial 12 bold",
            columnspan=2,
            sticky="w",
        )

        # Import button
        self.custom_button(
            text="Import List",
            row=1,
            column=2,
            font="Arial 12 bold",
            columnspan=1,
            width=15,
            sticky="nswe",
            command=self.import_button,
        )

        # Row 2
        # Select a student text label
        self.custom_text_label(
            text="Select a Student:",
            row=2,
            column=0,
            font="Arial 12 bold",
            columnspan=2,
            sticky="w",
        )

        # Section text label
        self.custom_text_label(
            text="Section:",
            row=2,
            column=2,
            font="Arial 12 bold",
            columnspan=1,
            sticky="nswe",
        )
        # Attended students text label
        self.custom_text_label(
            text="Attended Students:",
            row=2,
            column=3,
            font="Arial 12 bold",
            columnspan=2,
            sticky="w",
        )

        # Row 3
        # Section Student listbox
        self.update_listbox(index=0, side=0)

        # Section combobox
        self.update_combo_box()

        # Attended students listbox
        self.update_listbox(index=0, side=1)

        # Row 4
        # Add button
        self.custom_button(
            text="Add =>",
            row=4,
            column=2,
            font="Arial 10 bold",
            width=15,
            height=0,
            sticky="nswe",
            command=lambda: self.move_items_between_listboxes(
                self.list_left, self.list_right
            ),
        )

        # Row 5
        # Remove button
        self.custom_button(
            "<= Remove",
            row=5,
            column=2,
            width=15,
            height=0,
            font="Arial 10 bold",
            sticky="nswe",
            command=lambda: self.move_items_between_listboxes(
                self.list_right, self.list_left
            ),
        )

        # Row 6
        # Please selct file type text label
        self.custom_text_label(
            text=" Please select file type:",
            row=6,
            column=0,
            font="Arial 10 bold",
            width=5,
            sticky="nswe",
        )

        # File type combobox
        self.file_types = ["txt", "xls", "csv"]
        self.file_type_combo_box = self.custom_combo_box(
            row=6, column=1, width=5, height=5
        )
        self.file_type_combo_box["values"] = self.file_types
        self.file_type_combo_box.current(0)

        # Please enter week text label
        self.custom_text_label(
            text="Please enter week:",
            row=6,
            column=2,
            width=5,
            font="Arial 10 bold",
            sticky="nswe",
        )

        # Week input
        self.week_input = Entry(self.parent, width=15)
        self.week_input.grid(row=6, column=3, sticky="nswe", padx=3, pady=3)

        # Export as file button
        self.custom_button(
            text="Export as File",
            row=6,
            column=4,
            width=10,
            font="Arial 10 bold",
            sticky="nswe",
            command=self.export_button,
        )


def main():
    root = Tk()
    root.geometry("780x250")
    root.resizable(False, False)
    app = AttendanceKeeperApp(root)
    root.mainloop()


main()
